#!/usr/bin/env python3
"""
RMS Formblatt Service
HTTP-Server fuer PDF-Generierung aus XLSX-Templates

Laeuft auf Port 5050 und wird von n8n aufgerufen

Endpoints:
  POST /fill-xlsx     - Befuellt XLSX mit Daten
  POST /convert-pdf   - Konvertiert XLSX zu PDF
  POST /generate      - Kompletter Workflow (fill + convert)
  GET  /health        - Health Check

Autor: Claude Code
Datum: 2026-01-28
"""

import base64
import json
import os
import subprocess
import tempfile
import shutil
from pathlib import Path
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl nicht installiert")
    exit(1)

PORT = 5050
TEMP_DIR = Path("/tmp/rms-forms")
TEMP_DIR.mkdir(exist_ok=True)

# Feld-Mappings fuer verschiedene Formulare
FIELD_MAPPINGS = {
    "F_QM_02": {
        "I1": "abweichungs_nr",
        "I3": "datum",
        "A17": "lieferant_firma",
        "E23": "artikel_nr_schneider",
        "E24": "artikel_nr_lieferant",
        "E25": "artikel_bezeichnung",
        "E26": "lieferschein_nr",
        "E27": "lieferdatum",
        "E28": "liefermenge",
        "E29": "beanstandungsmenge",
        "A33": "beschreibung",
        "A47": "cb_untersuchung_abstellen",
        "A48": "cb_untersuchung_8d",
        "A49": "cb_ersatzlieferung",
        "A50": "cb_ruecksendung",
        "A51": "cb_gutschrift",
        "A52": "cb_sonstiges",
        "B56": "ersteller",
        "F56": "datum_unterschrift"
    },
    "F_QM_03": {
        "D3": "qa_id",
        "D4": "kunde",
        "D5": "artikel",
        "D6": "datum_reklamation",
        "B10": "d1_team",
        "B14": "d2_problembeschreibung",
        "B18": "d3_sofortmassnahme",
        "B22": "d4_ursachenanalyse",
        "B26": "d5_korrekturmassnahmen",
        "B30": "d6_wirksamkeit",
        "B34": "d7_vorbeugung",
        "B38": "d8_abschluss",
        "D40": "ersteller",
        "H40": "datum_abschluss"
    },
    "F_QM_04": {
        "D3": "nza_nr",
        "D4": "datum",
        "D5": "auftrag",
        "D6": "artikel",
        "B10": "fehler_beschreibung",
        "B14": "ursache",
        "B18": "massnahme",
        "D22": "aufwand_minuten",
        "D23": "kosten",
        "D26": "ersteller"
    },
    "F_QM_14": {
        "D3": "km_nr",
        "D4": "datum",
        "D5": "bezug_reklamation",
        "B10": "abweichung",
        "B14": "ursache",
        "B18": "korrekturmassnahme",
        "B22": "vorbeugung",
        "D26": "termin",
        "D27": "verantwortlich",
        "D30": "wirksamkeit_geprueft",
        "D31": "wirksamkeit_datum",
        "D34": "ersteller"
    }
}


def detect_form_type(filename: str) -> str:
    """Erkennt Formulartyp aus Dateinamen"""
    name = filename.upper()
    if "QM02" in name or "QM_02" in name:
        return "F_QM_02"
    elif "QM03" in name or "QM_03" in name or "8D" in name:
        return "F_QM_03"
    elif "QM04" in name or "QM_04" in name or "NZA" in name:
        return "F_QM_04"
    elif "QM14" in name or "QM_14" in name or "KORREKTUR" in name:
        return "F_QM_14"
    return None


def fill_xlsx(template_bytes: bytes, data: dict, form_type: str = None, filename: str = "template.xlsx") -> bytes:
    """Befuellt XLSX-Template mit Daten und gibt befuellte Datei zurueck"""

    # Template in temp-Datei schreiben
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
        tmp_in.write(template_bytes)
        input_path = tmp_in.name

    output_path = input_path.replace(".xlsx", "_filled.xlsx")

    try:
        # Formulartyp erkennen
        if not form_type:
            form_type = detect_form_type(filename)

        # Mapping holen
        if form_type and form_type in FIELD_MAPPINGS:
            mapping = FIELD_MAPPINGS[form_type]
        else:
            mapping = {}

        # Invertiertes Mapping
        field_to_cell = {v: k for k, v in mapping.items()}

        # Workbook laden und befuellen
        wb = load_workbook(input_path)
        ws = wb.active

        for field, value in data.items():
            if value is None or value == "":
                continue

            cell_ref = field_to_cell.get(field)
            if not cell_ref and len(field) <= 4 and field[0].isalpha():
                cell_ref = field

            if cell_ref:
                try:
                    if field.startswith("cb_"):
                        ws[cell_ref] = "☑" if value in [True, "true", "1", "ja", "yes"] else "☐"
                    else:
                        ws[cell_ref] = str(value)
                except Exception:
                    pass

        wb.save(output_path)
        wb.close()

        # Ergebnis lesen
        with open(output_path, "rb") as f:
            result = f.read()

        return result

    finally:
        # Cleanup
        if os.path.exists(input_path):
            os.remove(input_path)
        if os.path.exists(output_path):
            os.remove(output_path)


def convert_to_pdf(xlsx_bytes: bytes) -> bytes:
    """Konvertiert XLSX zu PDF mit LibreOffice"""

    # XLSX in temp-Datei schreiben
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(TEMP_DIR)) as tmp:
        tmp.write(xlsx_bytes)
        input_path = tmp.name

    try:
        # LibreOffice aufrufen
        env = os.environ.copy()
        env['HOME'] = '/tmp'

        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(TEMP_DIR), input_path],
            capture_output=True,
            text=True,
            timeout=120,
            env=env
        )

        if result.returncode != 0:
            raise Exception(f"LibreOffice error: {result.stderr}")

        # PDF-Pfad
        pdf_path = Path(input_path).with_suffix(".pdf")

        if not pdf_path.exists():
            raise Exception("PDF was not created")

        # PDF lesen
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

        # PDF loeschen
        pdf_path.unlink()

        return pdf_bytes

    finally:
        # Input loeschen
        if os.path.exists(input_path):
            os.remove(input_path)


class FormblattHandler(BaseHTTPRequestHandler):
    """HTTP Request Handler fuer Formblatt-Service"""

    def _send_response(self, status: int, data: dict):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def _send_binary(self, data: bytes, content_type: str, filename: str):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        if self.path == "/health":
            self._send_response(200, {"status": "ok", "service": "formblatt-service"})
        else:
            self._send_response(404, {"error": "Not found"})

    def do_POST(self):
        try:
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)
            data = json.loads(body) if body else {}

            if self.path == "/fill-xlsx":
                # Nur XLSX befuellen
                template_b64 = data.get("template")
                form_data = data.get("data", {})
                form_type = data.get("formType")
                filename = data.get("filename", "template.xlsx")

                if not template_b64:
                    self._send_response(400, {"error": "template required"})
                    return

                template_bytes = base64.b64decode(template_b64)
                filled_bytes = fill_xlsx(template_bytes, form_data, form_type, filename)

                self._send_response(200, {
                    "success": True,
                    "xlsx": base64.b64encode(filled_bytes).decode()
                })

            elif self.path == "/convert-pdf":
                # Nur zu PDF konvertieren
                xlsx_b64 = data.get("xlsx")

                if not xlsx_b64:
                    self._send_response(400, {"error": "xlsx required"})
                    return

                xlsx_bytes = base64.b64decode(xlsx_b64)
                pdf_bytes = convert_to_pdf(xlsx_bytes)

                self._send_response(200, {
                    "success": True,
                    "pdf": base64.b64encode(pdf_bytes).decode()
                })

            elif self.path == "/generate":
                # Kompletter Workflow: fill + convert
                template_b64 = data.get("template")
                form_data = data.get("data", {})
                form_type = data.get("formType")
                filename = data.get("filename", "template.xlsx")
                output_name = data.get("outputName", "formblatt")

                if not template_b64:
                    self._send_response(400, {"error": "template required"})
                    return

                # Template dekodieren
                template_bytes = base64.b64decode(template_b64)

                # XLSX befuellen
                filled_bytes = fill_xlsx(template_bytes, form_data, form_type, filename)

                # Zu PDF konvertieren
                pdf_bytes = convert_to_pdf(filled_bytes)

                self._send_response(200, {
                    "success": True,
                    "xlsx": base64.b64encode(filled_bytes).decode(),
                    "pdf": base64.b64encode(pdf_bytes).decode(),
                    "xlsxFilename": f"{output_name}.xlsx",
                    "pdfFilename": f"{output_name}.pdf"
                })

            else:
                self._send_response(404, {"error": "Not found"})

        except Exception as e:
            self._send_response(500, {"error": str(e)})

    def log_message(self, format, *args):
        print(f"[Formblatt-Service] {args[0]}")


def main():
    server = HTTPServer(("0.0.0.0", PORT), FormblattHandler)
    print(f"[Formblatt-Service] Listening on http://127.0.0.1:{PORT}")
    print(f"[Formblatt-Service] Endpoints: /health, /fill-xlsx, /convert-pdf, /generate")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Formblatt-Service] Shutting down...")
        server.shutdown()


if __name__ == "__main__":
    main()
