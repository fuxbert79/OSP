#!/usr/bin/env python3
"""
RMS Formular-Befueller
Befuellt XLSX-Templates mit Reklamationsdaten

Verwendung:
    python3 fill_xlsx_form.py template.xlsx output.xlsx --data '{"field": "value"}'
    python3 fill_xlsx_form.py template.xlsx --analyze  # Zeigt verfuegbare Felder

Autor: Claude Code
Datum: 2026-01-28
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl nicht installiert. Bitte ausfuehren:")
    print("  pip install openpyxl --break-system-packages")
    sys.exit(1)


# Feld-Mappings fuer verschiedene Formulare
FIELD_MAPPINGS = {
    "F_QM_02": {
        # Zelle: Datenfeld
        "I1": "abweichungs_nr",      # QA-ID
        "I3": "datum",                # Datum
        "A17": "lieferant_firma",     # Lieferant Firma (merged A17:E19)
        "E23": "artikel_nr_schneider",
        "E24": "artikel_nr_lieferant",
        "E25": "artikel_bezeichnung",
        "E26": "lieferschein_nr",
        "E27": "lieferdatum",
        "E28": "liefermenge",
        "E29": "beanstandungsmenge",
        "A33": "beschreibung",        # Beschreibung (merged A33:I42)
        # Checkboxen (als Text "☑" oder "☐")
        "A47": "cb_untersuchung_abstellen",
        "A48": "cb_untersuchung_8d",
        "A49": "cb_ersatzlieferung",
        "A50": "cb_ruecksendung",
        "A51": "cb_gutschrift",
        "A52": "cb_sonstiges",
        # Unterschrift
        "B56": "ersteller",
        "F56": "datum_unterschrift"
    },
    "F_QM_03": {
        "D3": "qa_id",
        "D4": "kunde",
        "D5": "artikel",
        "D6": "datum_reklamation",
        "B10": "d1_team",
        "B14": "d2_problembeschreibung",
        "B18": "d3_sofortmassnahme",
        "B22": "d4_ursachenanalyse",
        "B26": "d5_korrekturmassnahmen",
        "B30": "d6_wirksamkeit",
        "B34": "d7_vorbeugung",
        "B38": "d8_abschluss",
        "D40": "ersteller",
        "H40": "datum_abschluss"
    },
    "F_QM_04": {
        "D3": "nza_nr",
        "D4": "datum",
        "D5": "auftrag",
        "D6": "artikel",
        "B10": "fehler_beschreibung",
        "B14": "ursache",
        "B18": "massnahme",
        "D22": "aufwand_minuten",
        "D23": "kosten",
        "D26": "ersteller"
    },
    "F_QM_14": {
        "D3": "km_nr",
        "D4": "datum",
        "D5": "bezug_reklamation",
        "B10": "abweichung",
        "B14": "ursache",
        "B18": "korrekturmassnahme",
        "B22": "vorbeugung",
        "D26": "termin",
        "D27": "verantwortlich",
        "D30": "wirksamkeit_geprueft",
        "D31": "wirksamkeit_datum",
        "D34": "ersteller"
    }
}


def detect_form_type(filepath: str) -> str:
    """Erkennt Formulartyp aus Dateinamen"""
    filename = Path(filepath).stem.upper()

    if "QM02" in filename or "QM_02" in filename:
        return "F_QM_02"
    elif "QM03" in filename or "QM_03" in filename or "8D" in filename:
        return "F_QM_03"
    elif "QM04" in filename or "QM_04" in filename or "NZA" in filename:
        return "F_QM_04"
    elif "QM14" in filename or "QM_14" in filename or "KORREKTUR" in filename:
        return "F_QM_14"

    return None


def analyze_workbook(filepath: str):
    """Analysiert ein Workbook und zeigt Zellen mit Inhalt"""
    wb = load_workbook(filepath)

    print(f"\n=== Analyse: {filepath} ===")
    print(f"    Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n    Sheet: {sheet_name}")
        print(f"    Dimensionen: {ws.dimensions}")

        # Zeige Zellen mit Inhalt
        cells_with_content = []
        for row in ws.iter_rows(max_row=60, max_col=10):
            for cell in row:
                if cell.value:
                    cells_with_content.append(f"    {cell.coordinate}: {str(cell.value)[:50]}")

        print(f"    Zellen mit Inhalt ({len(cells_with_content)}):")
        for c in cells_with_content[:20]:
            print(c)
        if len(cells_with_content) > 20:
            print(f"    ... und {len(cells_with_content) - 20} weitere")

    wb.close()


def fill_workbook(template_path: str, output_path: str, data: dict, form_type: str = None):
    """Befuellt ein XLSX-Template mit Daten"""

    # Formulartyp erkennen
    if not form_type:
        form_type = detect_form_type(template_path)

    if not form_type or form_type not in FIELD_MAPPINGS:
        print(f"[WARN] Unbekannter Formulartyp. Verwende generisches Mapping.")
        # Generisches Mapping: Daten direkt in angegebene Zellen
        mapping = {k: k for k in data.keys() if len(k) <= 4}  # Nur Zell-Referenzen
    else:
        mapping = FIELD_MAPPINGS[form_type]
        print(f"[OK] Formulartyp erkannt: {form_type}")

    # Template laden
    wb = load_workbook(template_path)
    ws = wb.active

    # Invertiertes Mapping: Datenfeld -> Zelle
    field_to_cell = {v: k for k, v in mapping.items()}

    filled_count = 0

    for field, value in data.items():
        if value is None or value == "":
            continue

        # Zelle finden
        cell_ref = field_to_cell.get(field)

        if not cell_ref:
            # Vielleicht ist field direkt eine Zell-Referenz (z.B. "A1")
            if len(field) <= 4 and field[0].isalpha():
                cell_ref = field
            else:
                print(f"    [SKIP] Kein Mapping fuer Feld: {field}")
                continue

        try:
            # Checkbox-Handling
            if field.startswith("cb_"):
                ws[cell_ref] = "☑" if value in [True, "true", "1", "ja", "yes", "☑"] else "☐"
            else:
                ws[cell_ref] = str(value)

            filled_count += 1
            print(f"    [OK] {cell_ref} = {str(value)[:30]}")

        except Exception as e:
            print(f"    [ERR] Fehler bei {cell_ref}: {e}")

    # Speichern
    wb.save(output_path)
    wb.close()

    print(f"\n[OK] {filled_count} Felder befuellt")
    print(f"     Ausgabe: {output_path}")

    return filled_count > 0


def main():
    parser = argparse.ArgumentParser(description="RMS Formular-Befueller")
    parser.add_argument("template", help="Pfad zum XLSX-Template")
    parser.add_argument("output", nargs="?", help="Pfad zur Ausgabedatei")
    parser.add_argument("--data", "-d", help="JSON-String mit Felddaten")
    parser.add_argument("--data-file", "-f", help="JSON-Datei mit Felddaten")
    parser.add_argument("--analyze", "-a", action="store_true", help="Template analysieren")
    parser.add_argument("--form-type", "-t", help="Formulartyp (F_QM_02, F_QM_03, etc.)")

    args = parser.parse_args()

    # Template pruefen
    if not Path(args.template).exists():
        print(f"[ERR] Template nicht gefunden: {args.template}")
        sys.exit(1)

    # Analyse-Modus
    if args.analyze:
        analyze_workbook(args.template)
        return

    # Daten laden
    if args.data:
        try:
            data = json.loads(args.data)
        except json.JSONDecodeError as e:
            print(f"[ERR] Ungueltiges JSON: {e}")
            sys.exit(1)
    elif args.data_file:
        with open(args.data_file) as f:
            data = json.load(f)
    else:
        print("[ERR] Keine Daten angegeben. Verwende --data oder --data-file")
        sys.exit(1)

    # Output-Pfad
    if not args.output:
        template_path = Path(args.template)
        args.output = str(template_path.parent / f"filled_{template_path.name}")

    # Befuellen
    success = fill_workbook(args.template, args.output, data, args.form_type)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
